#
# The script's main function is to process pre-formatted DB data (RDBMS)
# and outputing an Excel (.xlsx) file.
#

import os
import sys
import csv
import json
import argparse
import shutil
import socket
import xlsxwriter
from decimal import *
import statistics
import numpy
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
import sqlalchemy
from sqlalchemy.orm import sessionmaker, Session
from sqlalchemy import MetaData, Table, Column, Integer, String, Float, create_engine, __version__
from sqlalchemy.ext.declarative import declarative_base

# constants
NUM_ARGS = 2
SOURCE_TAG="source"
XLSX_TAG="xlsx"
FNAME_TAG="fname"
PERF_DATA_SOURCE_TAG="perf_datasource"
DATA_SOURCE_CAT_DATABASE="database"
PRICING_DATA_SOURCE_TAG="pricing_datasource"
DATA_SOURCE_TAG="xlsx_report"
DB_SERVER_TAG="db_server"
HOST_ADDRESS_TAG="host"
USER_TAG="user"
PASSWORD_TAG="password"
DB_TAG="db"
PORT_TAG="port"
TABLE_TAG="table"
WORK_SHEET_TAG="worksheet"
NAME_TAG="name"
MYSQL_DEFAULT_PORT=3306

#report params
REPORT_PARAMS_TAG = "report"
OUTPUT_XLSX_FILE_PATH="output_xlsx_file"

# XLSX file tabs
XSLX_SERVERS_WS_NAME="Servers"
XSLX_PRICES_WS_NAME="Prices"
XSLX_RAW_WS_NAME="Raw"
XLSX_OVERALL_WS_NAME="Overall"

# Data table parameters
DT_UID_COL="uid"
DT_VM_CLASS_COL="vm_class"
DT_PROVIDER_COL="provider"
DT_RUNTIME_COL="runtime"

# end constants

# class GenUtils
class GenUtils():
    # getMin(self, inList)
    def getMin(self, inList):
        if inList is None: return None
        if len(inList) == 0: return None
        myVal = inList[0]
        
        for el in inList:
            if el < myVal: myVal = el

        return myVal
    # end get Min(self, list)

    # getMax(self, inList)
    def getMax(self, inList):
        if inList is None: return None
        if len(inList) == 0: return None
        myVal = inList[0]
        
        for el in inList:
            if el > myVal: myVal = el

        return myVal
    # end get Min(self, list)

    # convertListToDecimals(self, valList)
    def convertListToDecimals(self, valList):
        if valList is None: return []
        
        decList = []
        
        for v in valList:
            decList.append( Decimal(v) )

        return decList
    # end convertListToDecimals(self, valList)

    # convertListToFloats(self, valList)
    def convertListToFloats(self, valList):
        if valList is None: return []
        
        decList = []
        
        for v in valList:
            decList.append( float(v) )

        return decList
    # end convertListToFloats(self, valList)
    
# end class GenUtils

#
# Logging the messages.
#
# class Logger
class Logger():
    __logLevel = 0
    
    def logMessage(self, logLevel, message):
        if logLevel >= self.__logLevel:
            print message

    def setLogLevel(self, logLevel): self.__logLevel = logLevel

    def getLogLevel(self): return self.__logLevel
# end class Logger

# setting up logger
scriptLogger = Logger()
scriptLogger.setLogLevel(3)

# class TableTransaction
class TableTransaction():
    __tableData = None
    __index = None

    def __init__(self, tableData): self.__tableData = tableData

    # getColumn(self, columnName)
    def getColumn(self, columnName, ascendingIndex=None):
        fieldInex = None
        columnValueList = []
        runningIndex = ascendingIndex
        dataMatrix = self.__tableData.getTableMatrix()
        
        try:
            fieldIndex = self.__tableData.getColumnList().index( columnName )
        except:
            return columnValueList

        if runningIndex is None: runningIndex =\
           range(0,\
                 len( self.__tableData.getTableMatrix() ))

        rowCont = None
        
        for rowNum in runningIndex:
            rowCont = dataMatrix[rowNum]
            
            columnValueList.append( rowCont[fieldIndex] )
                
        return columnValueList
    # end getColumn(self, columnName)

    # getRowSelectionByIndex(self, index, ascendingFlag=True)
    def getRowSelectionByIndex(self, index, ascendingFlag=True):
        columnList = self.__tableData.getColumnList()
        dataMatrix= self.__tableData.getTableMatrix()
        selectionDataMatrix = []

        if index is None: return TableData(columnList, selectionDataMatrix)

        row_ref = None
        running_index = index
        
        for i in range(0, len(index)):
            if ascendingFlag:
                row_ref = running_index[0]
                if len(running_index) > 1: running_index = running_index[1:]
            else:
                row_ref = running_index[len(running_index)]
                if len(running_index) > 1: running_index =\
                   running_index[0:len(running_index)-1]
        
            selectionDataMatrix.append(dataMatrix[row_ref])
            
        return TableData(columnList, selectionDataMatrix)       
    # end getRowSelectionByIndex(self, index, ascendingFlag=True)

    # getRowIndexByFieldValue(self, fieldName, fieldValue)
    def getRowIndexByFieldValue(self, fieldName, fieldValue):
        selectionDataMatrix = []
        columnList = self.__tableData.getColumnList()
        dataMatrix = self.__tableData.getTableMatrix()
        selectionIndex = []

        try:
            fieldIndex = columnList.index()
        except:
            return selectionIndex

        for i in range(0, len(dataMatrix)):
            if (dataMatrix[i][fieldIndex] == fieldValue):
                selectionIndex.append(i)

        return selectionIndex      
    # getRowIndexByFieldValue(self, fieldName, fieldValue)

    # getSelectionByFieldValue(self, fieldName, fieldValue)
    def getSelectionByFieldValue(self, fieldName, fieldValue):
        columnList = self.__tableData.getColumnList()
        dataMatrix = self.__tableData.getTableMatrix()
        selectionMatrix = []
        selectionIndex = []

        try:
            fieldIndex = columnList.index(fieldName)
        except:
            return TableData(columnList, selectionMatrix)

        for i in range(0, len(dataMatrix)):
            if (dataMatrix[i][fieldIndex] == fieldValue):
                selectionMatrix.append(dataMatrix[i])
                
        return TableData(columnList, selectionMatrix)
    # end getSelectionByFieldValue(self, fieldName, fieldValue)

    # Appends table to the botom of the core one. Names and number of fields
    # have to match. Returns summary table.
    # getAppendedTable(self, table)
    def getAppendedTable(self, table):
        columnList = self.__tableData.getColumnList()
        summaryTableMatrix = []
        
        for i in range(0, len(self.__tableData.getTableMatrix())):
            summaryTableMatrix.append( (self.__tableData.getTableMatrix())[i] )

        for i in range(o, len(table.getTableMatrix()) ):
            summaryTableMatrix.append( (table.getTableMatrix())[i] )

        return TableData(columnList, summaryTableMatrix)
    # getAppendedTable(self, table)
    
# end class TableTransaction

# class TableData
class TableData():
    __tableMatrix = None;
    __columnList = None

    def __init__(self, columnList, tableMatrix):
        self.__columnList = columnList
        self.__tableMatrix = tableMatrix

    def getTableMatrix(self): return self.__tableMatrix

    def getColumnList(self): return self.__columnList
        
# end class TableData

# class UIDDecoded
class UIDDecoded():
    __provider = None
    __vmClass = None
    __fullUID = None

    # __init__(self, uid, provider)
    def __init__(self, uid, provider):
        self.__fullUID = uid
        self.__provider = provider
    
        if self.__provider is None: return

        cutUid = uid[len(self.__provider):]
        
        if cutUid[0] == "-":
            cutUid = cutUid[1:]
            
        self.__vmClass = str(cutUid).split('-')[0]
    # end __init__(self, uid, tableData = None)

    def getFullUID(self): return self.__fullUID

    def getVMClass(self): return self.__vmClass

    def getProvider(self): return self.__provider
# end class UIDDecoded

# class PerfDataHandler
class PerfDataHandler():
    __dbConnectParams = {}
    __dbEngine = None
    __dbSession = None
    __dbTable = None
    __metadata = MetaData()
    __inputFileHandler = None
    __perfTableData = None
    
    def __init__(self): pass

    # __init__(self, inputFileHandler)
    def __init__(self, inputFileHandler):
        self.__inputFileHandler = inputFileHandler
        self.__dbConnectParams =  inputFileHandler.getDBConnectParams()
        self.establishDBConnection()
        self.establishDBSession()
        self.establishDBTable()
        self.__populateDataMatrix()
    # __init__(self, dcConnParams)
    
    # __populateDataMatrix()
    def __populateDataMatrix(self):
        if self.__dbEngine is None: return

        if self.__metadata is None: return

        if self.__dbTable is None: return

        # Adding "vm_class" after "uid"
        colName = None
        columnList = []
        
        for col in self.__dbTable.columns:
            colName = str(col).split('.')[1]
            columnList.append( colName )
            if colName == DT_UID_COL: columnList.append(DT_VM_CLASS_COL)

        print "DEBUG: pt4: " + str(columnList)
        uidIndex = columnList.index(DT_UID_COL)
        provIndex = columnList.index(DT_PROVIDER_COL)
        vmClassIndex = columnList.index(DT_VM_CLASS_COL)
        tableMatrix = []
        vmClass = None
        row_list = None
        elem_cnt = None
        
        for row_cont in self.__dbSession.query(self.__dbTable):
            row_list = []

            elem_cnt = -1
            for elem in row_cont:
                row_list.append(str(elem))
                elem_cnt += 1
                
                if elem_cnt == uidIndex:            
                    row_list.append( "" )

            row_list[vmClassIndex] = UIDDecoded( row_list[uidIndex],\
                                                 row_list[provIndex] ).\
                                                 getVMClass()
                
            tableMatrix.append( row_list )

        self.__perfTableData = TableData( columnList, tableMatrix )
    # end __populateDataMatrix()   

    def setDBConnectParams(self, dbConnParams): self.__dbConnectParams = dbConnParams

    def getDBConnectParams(self): return self.__dbConnectParams
    
    # establishDBConnection(self)
    def establishDBConnection(self):
        if self.__dbConnectParams[DB_SERVER_TAG] is not None:
            self.__dbEngine = sqlalchemy.create_engine
            ("mysql://%s:%s@%s:%s/%s" %
             (self.__dbConnectParams[DB_SERVER_TAG][USER_TAG],
              self.__dbConnectParams[DB_SERVER_TAG][PASSWORD_TAG],
              self.__dbConnectParams[DB_SERVER_TAG][HOST_ADDRESS_TAG],
              self.__dbConnectParams[DB_SERVER_TAG][PORT_TAG],
              self.__dbConnectParams[DB_SERVER_TAG][DB_TAG]))            
    # end establishDBConnection(self)

    # establishDBSession(self)
    def establishDBSession(self):
        if self.__dbEngine is None: return
        
        self.__dbSession = Session(bind=self.__dbEngine )
    # end establishDBSession(self)
    
    # establishDBTable(self)
    def establishDBTable(self):
        if self.__dbEngine is None: return

        if self.__metadata is None: return

        if self.__dbConnectParams is None: return

        print "DEBUG: pt1: " + str( dir(self.__dbEngine) ) + \
            " | " + str( self.__dbEngine ) + \
            " | " + str( dir(self.__metadata) ) + \
            " | " + str( self.__metadata )
        
        self.__metadata.reflect(bind=self.__dbEngine)

        print "DEBUG: pt2: " + str(self.__metadata)
        
        self.__dbTable = Table(self.__dbConnectParams[DB_SERVER_TAG][TABLE_TAG],\
                               self.__metadata)    
    # end establishDBTable(self)
    
    def getDBConnection(self): return self.__dbEngine

    def getDBSession(self): return self.__dbSession

    def getDBTable(self): return self.__dbTable

    def getInputFileHandler(self): return self.__inputFileHandler

    def getPerfTableData(self): return self.__perfTableData

    def getUniqueSortedFieldValues(self, fieldName):
        columnValues = TableTransaction(self.__perfTableData).getColumn(fieldName)
        
        if columnValues is None: return []

        return sorted( set(columnValues) )
# end class PerfDataHandler

# class PriceDataHandler
class PriceDataHandler():
    __outputFileHandler = None
    __priceXLSXFilePath = None
    __priceDataMatrix = None

    # __init__(self, outputFileHandler)
    def __init__(self, outputFileHandler):
        self.__outputFileHandler = outputFileHandler
        self.__priceXLSXFilePath = self.__getPriceXLSXFilePath()
        self.__priceDataMatrix = self.__getPriceDataMatrix()
    # end __init__(self, outputFileHandler)
        
    # __getPriceXLSXFilePath(self)
    def __getPriceXLSXFilePath(self):
        filePath = None
        xlsxConfirmed = False
        xmlTree = self.__outputFileHandler.getInputFileHandler().getXMLTree()

        for member in xmlTree.getroot():
            if member.tag == PRICING_DATA_SOURCE_TAG:
                for subm in member:
                    if (subm.tag == SOURCE_TAG) and \
                       (subm.text == XLSX_TAG):
                        xlsxConfirmed = True

                    if (subm.tag == FNAME_TAG):
                        filePath = subm.text
                break
            
        if xlsxConfirmed:
            return filePath

        return None
    # end __getPriceXLSXFilePath(self)

    def getPriceXLSXFilePath(self): return self.__priceXLSXFilePath

    # __getPriceDataMatrix(self)
    def __getPriceDataMatrix(self):
        priceDataMatrix = []

        priceWorkBook = load_workbook(filename=self.__priceXLSXFilePath,\
                                      read_only=True)

        priceWs = priceWorkBook[XSLX_PRICES_WS_NAME]

        for row in priceWs.rows:
            rowList=[]
            for cell in row:
                rowList.append( cell.value )
            priceDataMatrix.append( rowList )

        return priceDataMatrix
    # end __getPriceDataMatrix(self)

    def getPriceDataMatrix(self): return self.__priceDataMatrix
    
# end class PriceDataHandler

# class InputFileHandler
class InputFileHandler():
    __filePath = None
    __paramsTree = None
    __dbConnectParams = {}
    __report_tree = None
    __pricing_datasource_tree = None
    __reportFilePath = None
    
    # __init__ (fname)
    def __init__(self, paramsFilePath = None):
        cnt1 = None
        cnt2 = None
        
        self.__filePath = paramsFilePath

        if self.__filePath is None: return
        
        self.__paramsTree = json.load( open(self.__filePath) )

        for member in self.__paramsTree:
            cnt1 = self.__paramsTree[member]
            
            if member == PERF_DATA_SOURCE_TAG:
                cnt2 = self.__paramsTree[member]

                for member2 in cnt2:
                    if member2 == DATA_SOURCE_CAT_DATABASE:
                        self.__dbConnectParams = cnt2[member2]

            if member == DATA_SOURCE_TAG:
                self.__report_tree = self.__paramsTree[member]

            if member == PRICING_DATA_SOURCE_TAG:
                self.__pricing_datasource_tree = self.__paramsTree[member]

            if member == REPORT_PARAMS_TAG:
                cnt2 = self.__paramsTree[member]

                for member1 in cnt2:
                    if member1 == OUTPUT_XLSX_FILE_PATH:
                        self.__reportFilePath = cnt2[member1]
                
    # end __init__ (fname)

    def getFilePath(self): return self.__filePath

    def getParamsTree(self): return self.__paramsTree

    def getDBConnectParams(self): return self.__dbConnectParams

    def getReportDataTree(self): return self.__report_tree

    def getPricingDataSourceTree(self): return self.__pricing_datasource_tree

    def getReportFilePath(self): return self.__reportFilePath
# end class InputHandler

# class OutputFileHandler
class OutputFileHandler():
    __argHandler = None
    __inputFileHandler = None
    __filePath = None
    __workbook = None
    __perfDataHandler = None
    __rawDataTable = None
    __dbSession = None
    __xlsx_report_tree = None
    
    # WS formating constants
    __raw_tab_name = "Raw"
    
    # __init__(self, argHandler, inputFileHandler)
    def __init__(self, inputFileHandler):
        self.__inputFileHandler = inputFileHandler
        self.__filePath = self.__inputFileHandler.getReportFilePath()
        self.__perfDataHandler = PerfDataHandler(inputFileHandler)
        self.__initWorkbook(self.__filePath)

        # Removing output file if it exists
        if os.path.isfile( self.__filePath ):
            os.remove( self.__filePath )
            
        xlsx_report_tree = inputFileHandler.getXLSXReportTree()
        dbTable = self.__perfDataHandler.getDBTable()
        dbSession = self.__perfDataHandler.getDBSession()

        if xlsx_report_tree is not None:
            self.setReportTree( xlsx_report_tree )

        if dbSession is not None:
            self.setDBSession( dbSession)
            
        if dbTable is not None:
            self.setRawDataTable( dbTable )

        self.__initWorksheets()
    # end __init__(self, fname, xlsx_report_tree, dbTable)

    # __initWorkbook(self, fname)
    def __initWorkbook(self, fname):
        self.__workbook = xlsxwriter.Workbook( fname )
    # end __initWorkbook(self, fname)

    # __initWorksheets(self)
    def __initWorksheets(self):        
        if self.__workbook is None: return

        if self.__xlsx_report_tree is None: return

        for member in self.__xlsx_report_tree:
            if member.tag == WORK_SHEET_TAG:
                for subm in member:
                    if subm.tag == NAME_TAG:
                        self.__workbook.add_worksheet(subm.text)
                        break
    # end __initWorksheets(self)

    def getArgHandler(self): return self.__argHandler

    def getInputFileHandler(self): return self.__inputFileHandler
    
    def setXLSXReportTree(self, xlsx_report_tree):
        self.__xlsx_report_tree = xlsx_report_tree

    def getFilePath(self): return self.__filePath

    def getWorkBook(self): return self.__workbook

    def setRawDataTable(self, dbTable): self.__rawDataTable = dbTable

    def getRawDataTable(self): return self.__rawDataTable

    def setDBSession(self, dbSession): self.__dbSession = dbSession

    def getDBSession(self): return self.__dbSession

    def getPerfDataHandler(self): return self.__perfDataHandler

    # fillRawTab(self)
    def fillRawTab(self):        
        if (self.__rawDataTable is None) or \
           (self.__dbSession is None): return
        
        ws = self.__workbook.get_worksheet_by_name(XSLX_RAW_WS_NAME)

        if ws is None: return

        c_count = 0
        for col in self.__rawDataTable.columns:
            ws.write(0, c_count, str(col).split('.')[1] )
            c_count += 1

        r_count = 1
        for row_cont in self.__dbSession.query(self.__rawDataTable):
            c_count = 0
            for elem in row_cont:
                ws.write(r_count, c_count, elem)
                c_count += 1
            r_count += 1
    # end fillRawTab(self)

    # fillPricesTab(self)
    def fillPricesTab(self):
        ws = self.__workbook.get_worksheet_by_name(XSLX_PRICES_WS_NAME)

        if ws is None: return

        pdh = PriceDataHandler(self)

        pdm = pdh.getPriceDataMatrix()

        if pdm is None: return

        r_count = 0
        for row in pdm:
            c_count = 0
            for elem in row:
                ws.write(r_count, c_count, elem)
                c_count += 1
            r_count+=1
    # end  fillPricesTab(self)

    # fillOverallTab(self)
    def fillOverallTab(self):
        if self.__perfDataHandler is None: return
        
        ws = self.__workbook.get_worksheet_by_name(XLSX_OVERALL_WS_NAME)

        if ws is None: return

        tableData = self.__perfDataHandler.getPerfTableData()

        if tableData is None: return

        columnList = tableData.getColumnList()
        tableMatrix = tableData.getTableMatrix()

        provIndex = columnList.index( DT_PROVIDER_COL )
        vmClassIndex = columnList.index( DT_VM_CLASS_COL )
        runtimeIndex = columnList.index( DT_RUNTIME_COL )

        providerList = sorted(set(TableTransaction(tableData).getColumn(DT_PROVIDER_COL)))

        tableMatrixSel = []
        vmClass = None
        vmClassList = None
        prvDataTable = None
        runtimeDataList = None
        intmultiDataList = None
        memmultiDataList = None

        xlsRowCnt = 1

        ws.write(0, 0, "Provider")
        ws.write(0, 1, "VM Class")
        ws.write(0, 2, "Runtime min")
        ws.write(0, 3, "Runtime max")
        ws.write(0, 4, "Runtime stdev")
        ws.write(0, 5, "Runtime 5th percentile")
        ws.write(0, 6, "Runtime 95th percentile")
        
        for prv in providerList:
            if prv is None: continue
            
            prvDataTable = TableTransaction(tableData).\
                           getSelectionByFieldValue(DT_PROVIDER_COL, prv)
          
            vmClassList = sorted(set(TableTransaction(prvDataTable).\
                                     getColumn(DT_VM_CLASS_COL)))

            
            for vmClass in vmClassList:
                if vmClass is None: continue
                
                scriptLogger.logMessage(5, prv + " " + vmClass )
                
                prvVMClassTable = TableTransaction(prvDataTable).\
                                  getSelectionByFieldValue(DT_VM_CLASS_COL, vmClass)

                runtimeDataList = TableTransaction(\
                                                   prvVMClassTable).\
                                                   getColumn("runtime")

                scriptLogger.logMessage(5, "Runtime list " +\
                                        str(len(runtimeDataList)) +\
                                        " elements")

                try:
                    runtimeDataList = GenUtils().\
                                      convertListToFloats(runtimeDataList)

                 
                    intmultiDataList = GenUtils().\
                                       convertListToFloats(TableTransaction(prvVMClassTable).\
                                                           getColumn("intmulti"))

                    memmultiDataList = GenUtils().\
                                       convertListToFloats(TableTransaction(prvVMClassTable).\
                                                           getColumn("memmulti"))
                except:
                    scriptLogger.logMessage(5, "WARNING: bad data")
                    continue
                    
                runtimeMin = GenUtils().getMin(runtimeDataList)

                runtimeMax = GenUtils().getMax(runtimeDataList)

                runtimeMedian = statistics.median(runtimeDataList)

                runtimeMean = statistics.mean(runtimeDataList)

                try:
                    runtimeStDev = statistics.stdev(runtimeDataList)
                except:
                    runtimeStDev = 0
                    
                runtime5thPerc = numpy.percentile(runtimeDataList, 5)

                runtime95thPerc = numpy.percentile(runtimeDataList, 95)

                ws.write(xlsRowCnt, 0, prv)
                ws.write(xlsRowCnt, 1, vmClass)
                ws.write(xlsRowCnt, 2, runtimeMin)
                ws.write(xlsRowCnt, 3, runtimeMax)
                ws.write(xlsRowCnt, 4, runtimeMedian)
                ws.write(xlsRowCnt, 5, runtimeMean)
                ws.write(xlsRowCnt, 6, runtimeStDev)
                ws.write(xlsRowCnt, 7, runtime5thPerc)
                ws.write(xlsRowCnt, 8, runtime95thPerc)

                xlsRowCnt += 1
                
    # end fillOverallTab(self)
    
    # commitWorkBook(self):
    def commitWorkBook(self):
        self.__workbook.close()
        self.__workbook = None
    # end commitWorkBook(self):
        
# end class OutputFileHandler

# class ReportGenerator
class ReportGenerator():
    __paramsFilePath = None
    __inputFileHandler = None
    __outputFileHandler = None

    # __init__(self, paramsFilePath = None)
    def __init__(self, paramsFilePath = None):
        self.__paramsFilePath = paramsFilePath

        if (self.__paramsFilePath is None):
            if ( len(sys.argv) < 2 ): return
            self.__paramsFilePath = sys.argv[1]
            
        self.__inputFileHandler = InputFileHandler(self.__paramsFilePath)
        self.__outputFileHandler = OutputFileHandler(self.__inputFileHandler)
        
    # end __init__(self, paramsFilePath = None)

    # runXLSXReport(self)
    def runStdXLSXReport(self):
        if (self.__outputFileHandler is None): return

        self.__outputFileHandler.fillRawTab()
        self.__outputFileHandler.fillOverallTab()
    # end runXLSXReport(self)
        
# end class ReportGenerator

# main block invocation
if __name__ == "__main__":
    if ( len(sys.argv) == 1 ): exit(0)
    
    repGen = ReportGenerator( sys.argv[1] )
    repGen.runStdXLSXReport()
# end main block invocation
